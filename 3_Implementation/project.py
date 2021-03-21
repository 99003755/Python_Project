# -*- coding: utf-8 -*-
"""
Created on Tue Mar 16 06:25:50 2021

@author: Rohan Roy
"""
from openpyxl import load_workbook
from openpyxl.styles import Font

path = "D:\Python_Practice\LnT.xlsx"
wb = load_workbook(path)

# Entering user defined data

name = input("Enter Name: ")
PS = eval(input("Enter PS Number: "))
email = input(" Enter email: ")


head = []

# Class for reading writing and printing the data to excel sheet


class Excel:

    def __init__(self, s_name, s_PS, s_email, s_path):
        self.name = s_name
        self.PS = s_PS
        self.email = s_email
        self.path = s_path

# Function to search data present in the excel sheet
    def data_search(self):
        s_wb = load_workbook(self.path)
        found1 = 0
        for sheet in s_wb.sheetnames:  # traversing through all the sheets
            print("IN SHEET")
            ws = s_wb[sheet]
            s = ws.max_row  # variable to store max rows for sl num
            col = ws.max_column
            for i in range(1, s + 1):
                if ws.cell(row=i, column=1).value == self.name and ws.cell(row=i, column=2).value == self.PS \
                        and ws.cell(row=i, column=3).value == self.email: # searching data provided by user
                    if sheet == "Sheet0":
                        break
                    if sheet != 'Sheet1': 

                        for k in range(4, col+1):
                            data.append(ws.cell(row=i, column=k).value)  # appending data to put in master sheet
                            head.append(ws.cell(row=1, column=k).value)  # appending head to put in master sheet
                    else:
                        for j in range(1, col+1):
                            # print(ws.cell(row=i, column=j).value)
                            data.append(ws.cell(row=i, column=j).value)
                            head.append(ws.cell(row=1, column=j).value)
                            # Variable to check if data is present in excel sheet
                    found1 = 1

            print(data)
        return data, found1

# function to write the data in the master sheet
    def data_write(self, w_data, w_found):
        e_data = w_data
        e_found = w_found
        w_wb = load_workbook(self.path)
        if e_found == 1:
            if 'Sheet0' not in wb.sheetnames:
                ws = w_wb.create_sheet('Sheet0')
                print("CREATING")
                s = ws.max_row  # variable to store max rows for sl num
                for i in range(1, len(head)+1):
                    ws.cell(row=1, column=i).value = head[i - 1]  # Add headings to sheet 
                for i in range(1, len(head)+1):
                    clr = ws.cell(row=1, column=i)
                    clr.font = Font(bold=True)  # adding font to headings on the sheet
                for i in range(1, len(head)+1):
                    ws.cell(row=s + 1, column=i).value = e_data[i - 1]
                w_wb.save(self.path)
            else:
                # ws = wb.get_sheet_by_name('Sheet0')
                ws = wb['Sheet0']
                s = ws.max_row
                for i in range(1, len(head)+1):
                    ws.cell(row=s + 1, column=i).value = e_data[i - 1]
                w_wb.save(self.path)
        if e_found == 0:
            print("DATA NOT FOUND")

# function to display the bar chart on the excel sheet
    def bar(self):
        from openpyxl.chart import BarChart3D, Reference
        b_wb = load_workbook(self.path)
        ws = b_wb['Sheet0']

        chart1 = BarChart3D()
        # adding title, x-axis and y-axis to the bar chart
        chart1.title = "EXCEL DATA"  
        chart1.y_axis.title = 'Marks'  
        chart1.x_axis.title = 'Student'
        bar_r = ws.max_row
        bar_c = ws.max_column
        if bar_r <= 2:
            bar_r = 3
        e_data = Reference(ws, min_col=4, min_row=abs(bar_r-2), max_row=bar_r, max_col=bar_c)

        chart1.add_data(e_data, titles_from_data=True)
        ws.add_chart(chart1, "J15")
        wb.save(self.path)

# Calling functions in the class


write = Excel(name, PS, email, path)
data, found = write.data_search()
write.data_write(data, found)
write.bar()
